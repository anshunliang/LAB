import sys,time,socket
from openpyxl import *
import random,queue,openpyxl,cgitb,threading
from functools import partial
from PyQt5.QtCore import *
from matplotlib import pyplot as plt
from scipy.interpolate import make_interp_spline
from PyQt5.QtWidgets import *
from PyQt5 import *
from ctypes import *
from influxdb import InfluxDBClient
from auto import autocontronl

from globalfile import globalvar as gl   #导入自建的全局变量，用存储所有工位号及其实时数据，内部时字典格式
from globalfile import gwh as gw         #导入工位号全局变量
from globalfile import shuju as sj        #导入数据全局变量

from x import Ui_MainWindow   #导入主窗口
from sub import datab
from sub import booll,floatt    #导入开关阀和调节阀子窗口 
import tx,qx
 
gl._init()            #全局变量初始化
cgitb.enable()        #异常捕捉

lock = QMutex()   #实例化锁,针对数据读取和数据刷新两个线程
lock1 = QMutex()   #实例化锁,针对自建的全局变量
q=queue.Queue()   #实例化队列，针对控制消息获取


#建立全局变量
xp=[]  #用于存储AI控件
xpindex=[]
xpd=[]  #用于存储DI控件
xpdindex=[]
xptype=[]
xh=[]        #用于存储从TCP信号（未转换）
cs=[]        #用于存储信号（已转换）

jp=[]        #存储控件对应的数值索引
address=""   #用于存储TCP服务器地址及端口
size=0       #用于存储从TCP服务器接收的数据字节数
sss={}       #存储接收的原始数据
ssss={}      #存储数据

ts=0    #标记值，用于判定是否开始存储数据

'''
class Main(Ui_MainWindow):
    def __init__(self):
        super().__init__()
        
    def prt(self):
        print("jjjj")
'''

#建立TCP通信
tcp_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
server_addr = ("127.0.0.1",8000)
tcp_socket.connect(server_addr)
        



#读取配置表
data = openpyxl.load_workbook('AI.xlsx')
wsai = data.get_sheet_by_name('AI')
wsao = data.get_sheet_by_name('AO')
wsdi = data.get_sheet_by_name('DI')
wsdo = data.get_sheet_by_name('DO')

wsai_nrows = wsai.max_row  #获得AI行数
wsao_nrows = wsao.max_row  #获得AO行数
wsdi_nrows = wsdi.max_row  #获得DI行数
wsdo_nrows = wsdo.max_row  #获得DI行数

PZAI=[]                    #存储AI工位号
PZAI_ysll=[]               #存储AI工位号的原始下限值
PZAI_yshh=[]               #存储AI工位号的原始上限值
PZAI_ll=[]                 #存储AI工位号的工程下限值
PZAI_hh=[]                 #存储AI工位号的工程上限值

PZAO=[]                    #存储AO工位号
PZDI=[]
PZDO=[]

#读AI表格
wsai.cell(row=4, column=2, value=10)
for i in range(1,wsai_nrows):
    j= wsai.cell(row=i+1, column=1).value    #获取1+1行，第一列的值，工位号
    PZAI.append(j) 
    cdc=wsai.cell(row=i+1, column=3).value   #获取1+1行，第3列的值，原始值下限
    PZAI_ysll.append(cdc)                           
    cde=wsai.cell(row=i+1, column=4).value   #获取1+1行，第4列的值，原始值上限
    PZAI_yshh.append(cde) 
    cde=wsai.cell(row=i+1, column=5).value   #获取1+1行，第5列的值，工程值下限
    PZAI_ll.append(cde)                           
    cdf=wsai.cell(row=i+1, column=6).value   #获取1+1行，第6列的值，工程值上限
    PZAI_hh.append(cdf)


#读AO表格
for i in range(1,wsao_nrows):
    j= wsao.cell(row=i+1, column=1).value    #获取1+1行，第一列的值
    PZAO.append(j)


#读DI表格
for i in range(1,wsdi_nrows):
    j= wsdi.cell(row=i+1, column=1).value    #获取1+1行，第一列的值
    PZDI.append(j)


#读Do表格
for i in range(1,wsdo_nrows):
    j= wsdo.cell(row=i+1, column=1).value    #获取1+1行，第一列的值
    PZDO.append(j)

#计算模拟量的个数
xo=len(PZAI)+len(PZAO)
gw.set(PZAI+PZAO+PZDI+PZDO)


#定义C指针

print(len(PZAI))
print(PZAI_ysll)
pv = (c_float*len(PZAI))(*PZAI_ysll)       #模拟原始值,随便找一个浮点数组
ysxx = (c_float*len(PZAI))(*PZAI_ysll)     #模拟原始值下限
ysss = (c_float*len(PZAI))(*PZAI_yshh)     #模拟原始值上限
gzxx = (c_float*len(PZAI))(*PZAI_ll)         #模拟工程值下限
gzss = (c_float*len(PZAI))(*PZAI_hh)         #模拟工程值上限
xsxs=PZAI
xsxs = (c_float*len(PZAI))(*PZAI_hh)   #模拟换算结果，占位

foo=cdll.LoadLibrary("F:\\jiemian\\C\\x1.dll")
foo.pp.argtypes = [POINTER(c_float),POINTER(c_float),POINTER(c_float),POINTER(c_float),POINTER(c_float),POINTER(c_float),c_int] 

ps=pointer(pv)
p1=pointer(ysxx)
p2=pointer(ysss)
p3=pointer(gzxx)
p4=pointer(gzss)


#foo.pp(pv,ysxx,ysss,gzxx,gzss,carrayjs)






#定义子窗口
class Child(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("使用教程")
        self.resize(250, 180)

        self.dk= QLabel(self)
        self.dk.setGeometry(30, 30, 100, 100)
        str='端口'
        self.dk.setText(str)
        
        
    def open(self):
        self.show()



#首次运行时，根据配置表搜索前面板控件的函数，
def click_success():
    #print("start qt5")
    
    global xp
    global PZAI
    global PZDI
    
    j1=0
    for i in PZAI:
        try:
            if MainWindow.findChild(QLineEdit,i):   #根据反馈AI工位号获取前面板控件
               xp.append(MainWindow.findChild(QLineEdit,i))
               xpindex.append(j1)
               j1+=1
            else:
                j1+=1
        except:
            pass
    j1=0
    j=0
  
    for i in PZDI:                                 #根据反馈DI工位号获取前面板控件
        try:
            if MainWindow.findChild(QPushButton,i):
                xpd.append(MainWindow.findChild(QPushButton,i))
                xpdindex.append(j)
                j+=1
            else:
                j+=1
        except:
            pass


#读取信号线程
class Mythread0(QThread):
    # 定义信号,定义参数为str类型
    breakSignal = pyqtSignal(object)
 
    def __init__(self, parent=None):
        super().__init__(parent)

    def run(self):
        time.sleep(1)
        global cs
        global xp
        global xh
        global tcp_socket
        while True:
            time.sleep(0.1)
           
            #发送信号到刷新函数,触发界面刷新函数
            #self.breakSignal.emit(str(a))
            #模拟产生现场信号,用于刷新主界面
            lock.lock()
            

            xxx=tcp_socket.recv(186)
            xhh=xxx.decode('UTF-8').split(" ")
            xh=xhh[1:]
            #sj.set(xh)
            #print(sj.get()[2])
            #将读取的信号写入自建全局变量
            try:
                j=0
                for i in PZAI+PZAO+PZDI+PZDO:
                    gl.set_value(i, str(xh[j]))     #设定值
                    ssss[i]=str(xh[j])
                    j+=1
            except:
                pass
            
            zaq=0   #临时变量
            global ps
            for i in xh[0:len(PZAI)]:
                #cccc.append(float(i))
                ps.contents[zaq]=float(i)
                print(ps.contents)
                zaq+=1
            zaq=0
            #调用动态链接库，执行换算
            foo.pp(pv,ysxx,ysss,gzxx,gzss,xsxs,len(PZAI))

            global cs
            cs=[]
            for i in xsxs:
                cs.append(round(i, 2))
            #print(cs)
            sj.set(cs)
            
            
            
            #发送信号到刷新函数,触发界面刷新函数
            self.breakSignal.emit(xh)


            lock.unlock()
            
            


#
class Mythread1(QThread):

    # 定义信号,定义参数为int类型
    def __init__(self, parent=None):
        super().__init__(parent)
    
    def run(self):
        time.sleep(2)
        while True:
            time.sleep(0.1)


#获取操作员命令，发送控制信号的线程,将控制信号发送到下位机
class Mythread2(QThread):
    # 定义信号,定义参数为int类型
    breakSigna2 = pyqtSignal(int)
    def __init__(self, parent=None):
        super().__init__(parent)

    def run(self): 
        while True:
            time.sleep(0.1)
            #此处获取操作员命令 
            if not q.empty():
                xxx=q.get()
                ui.ee.setText(str(xxx))
                send_data = xxx
                tcp_socket.send(send_data.encode("utf-8"))

#界面刷新函数，a是线程发送过来的信号，用于刷新界面
def chuli(a):
    global xh
    global xpindex   #模拟量工位号索引
    global xpdindex  #布尔量工位号索引
    global xp      #模拟量工位号
    global xpd     #布尔量工位号
    global xo
    global cs
    autocontronl(q)   #执行自动控制代码
    
    for i,j in zip(xp,xpindex):
        #i.setText(str(xh[j]))
        i.setText(str(cs[j]))

              
    
    for k,p in zip(xpd,xpdindex):
        if str(xh[xo+p])=="1":
            k.setStyleSheet("background: rgb(0,255,0)")
        else:
            k.setStyleSheet("background: rgb(255,255,255)")
    

    

    
#开辟线程用于存储数据
class Mythreadx(QThread):
    # 定义信号,定义参数为int类型
    def __init__(self, parent=None):
        super().__init__(parent)
    
    def run(self):
        global xh
        global PZAI
        d={"a":"1","b":"2","c":"33333"}
        a=0
        client = InfluxDBClient('10.24.10.102', 8086,'root','123456',database='t1')
        while True:
            start=time.clock()
            time.sleep(0.1)
            
            localtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
            w_json = [{
                "measurement": 't16',
                "time": a,
                "tags": {
                'name': 1,
                'categories': 2
                    },
                "fields": ssss
            }]
            client.write_points(w_json)
            end=time.clock()
            print(str(end-start)+"秒")
           
            a=a+1
            ui.pc.setText(str(a))
            if ts==0:
                client.close()
                break
    
#打开 调节阀窗口函数
def convert( x,p):
    p.open()

    
#打开 开关阀窗口函数
def convertt( x,p):
    p.open()



###不能在函数内部实例化线程类
threadx = Mythreadx()
#存储函数
def save():
    global ts
    if ts==0:
        ts=1
        ui.p4.setStyleSheet('''QPushButton{background:#32CD32;border-radius:5px;}QPushButton:hover{background:green;}''')
        ui.p4.setText("正在存储")
        global threadx
        threadx.start()
    else:
        if ts==1:
            ts=0
            ui.p4.setStyleSheet('''QPushButton{background:#E6E6FA;border-radius:5px;}QPushButton:hover{background:green;}''')
            ui.p4.setText("开始存储")
    

    print("save")
def chhh():
    pass


#修改配置表时，查找工位号
def sr():
    #global PZAI_hh
    g11=ui.g1.text()
    if g11 in PZAI:
        s=PZAI.index(g11)
        ui.yuanshixx.setText(str(PZAI_ysll[s]))
        ui.yuanshisx.setText(str(PZAI_yshh[s]))
        ui.gcxx.setText(str(PZAI_ll[s]))
        ui.gcsx.setText(str(PZAI_hh[s]))
    else:
        ui.yuanshixx.setText("破剑式")
        ui.yuanshisx.setText("破刀式")
        ui.gcxx.setText("破枪式")
        ui.gcsx.setText("破剑式")


#修改配置表
def pzxgdef():
    
    

    g11=ui.g1.text()         #获取被改工位号
    if g11 in PZAI:
        #修改配置文件
        wb=load_workbook("AI.xlsx")
        ws=wb["AI"]
        s=PZAI.index(g11)
        PZAI_ll[s]=float(ui.gcxx.text())
        PZAI_hh[s]=float(ui.gcsx.text())
        ws.cell(s+2,3).value=float(ui.yuanshixx.text())
        ws.cell(s+2,4).value=float(ui.yuanshisx.text())
        ws.cell(s+2,5).value=float(ui.gcxx.text())
        ws.cell(s+2,6).value=float(ui.gcsx.text())
        wb.save("AI.xlsx")
        global p1,p2,p3,p4
        p1.contents[s]=float(ui.yuanshixx.text())
        p2.contents[s]=float(ui.yuanshisx.text())
        p3.contents[s]=float(ui.gcxx.text())
        p4.contents[s]=float(ui.gcsx.text())
        #修改配置文件



    
    
    


##继承全局数据窗口类
class MyWindow(QMainWindow, datab.Ui_MainWindow):
    def __init__(self, parent=None):
        super(MyWindow, self).__init__(parent)
        self.setupUi(self)
 

###继承主界面的类
class Main(Ui_MainWindow):
    def __init__(self):
        super().__init__()
    def prt(self):
        pass


if __name__ == '__main__':
    
    app = QApplication(sys.argv)
    MainWindow = QMainWindow()
    ui = Main()
    ui.setupUi(MainWindow)
    
    #第一种打开子窗口的方式
    #实例化自定义窗口
    #实例化自定义子窗口
    child = QMainWindow()          
    child_ui = tx.Ui_MainWindow()
    child_ui.setupUi(child)
    ui.tx.clicked.connect( child.show ) 
  
    #打开全局数据窗口
    #打开QTdesinger画的子窗口，并在子窗口里面重写了open事件
    m=MyWindow()
    ui.pzb.clicked.connect( m.open) 

    #打开曲线窗口
    '''
    cqx = QMainWindow()          
    cqx_ui = qx.Ui_MainWindow()
    cqx_ui.setupUi(cqx)
    ui.opqx.clicked.connect( cqx.show )   #打开
    '''
    
    
   
    #第二种打开子窗口的方式
    #实例化阀门窗口
    eh=floatt.Demo("e2",q,lock1)         #"e2代表阀门号。q代表传入子窗口的队列，用于发送控制消息。lock1用于锁住自建全局变量"
    dh=booll.Demo("e1",q,lock1)
    #打开子窗口
    ui.p1.clicked.connect(dh.open)
    ui.p2.clicked.connect(eh.open)


    #第三种打开子窗口的方式
    #重新实例化子窗口
    #Ph=floatt.Demo("e2",q,lock1)         #"e2代表阀门号。q代表传入子窗口的队列，用于发送控制消息。lock1用于锁住自建全局变量"
    #Ph=floatt.Demo()

    #为p1控件（阀门按钮）绑定事件
    ui.p5.clicked.connect(partial(convert, "p5",floatt.Demo("p5",q,lock1)))
    ui.p6.clicked.connect(partial(convert, "p6",floatt.Demo("p6",q,lock1)))
    ui.p7.clicked.connect(partial(convert, "p7",floatt.Demo("p7",q,lock1)))

    ui.bb5.clicked.connect(partial(convertt, "b5",booll.Demo("b5",q,lock1)))
    ui.bb6.clicked.connect(partial(convertt, "b6",booll.Demo("b6",q,lock1)))
    ui.bb7.clicked.connect(partial(convertt, "b7",booll.Demo("b7",q,lock1)))

    #修改保存按钮外观
    ui.p4.setStyleSheet('''QPushButton{background:#E6E6FA;border-radius:5px;}QPushButton:hover{background:green;}''')
    ui.p4.setText("开始存储")

    #点击存储按钮
    #ui.p4.clicked.connect(partial(save))
    ui.p4.clicked.connect(save)
    #首次运行，根据配置表搜索前面板控件的函数
    click_success()
   
    # 创建线程,用于读取信号，自动后台运行
    thread0 = Mythread0()
    thread0.breakSignal.connect(chuli)  
    thread0.start()

    # 创建线程,用于刷新界面信号，自动后台运行
    thread1 = Mythread1()
    thread1.start()

    #创建线程，用于发送控制信号
    thread2 = Mythread2()
    thread2.breakSigna2.connect(chuli)
    thread2.start()
    
    ####
    #配置表修改
    #监听提取间隔输入事件
    ui.g1.textChanged.connect(sr)
    ui.pzxg.clicked.connect(pzxgdef)
  
    MainWindow.show()
    sys.exit(app.exec_())